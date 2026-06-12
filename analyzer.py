"""
CSV Insight Analyzer
====================
Herramienta de análisis de datos empresarial para exploración,
estadística descriptiva y reporte automático de archivos CSV.

Autor  : MillesvcStudio
Versión: 1.0.0
"""

import sys
import warnings
from pathlib import Path
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use("Agg")                     # backend sin pantalla (compatible con servidores)
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────

REPORTS_DIR = Path("reports")
DATA_DIR    = Path("data")

PALETTE = {
    "primary"   : "#1A1A2E",
    "accent"    : "#E94560",
    "secondary" : "#16213E",
    "muted"     : "#9EB8D9",
    "bg"        : "#0F3460",
    "white"     : "#FFFFFF",
}

SECTION_SEP = "─" * 65


# ─────────────────────────────────────────────
# MÓDULO: CARGA DE DATOS
# ─────────────────────────────────────────────

def load_csv(file_path: str | Path) -> pd.DataFrame:
    """
    Carga un archivo CSV y devuelve un DataFrame de pandas.

    Args:
        file_path: Ruta al archivo CSV (string o Path).

    Returns:
        DataFrame con los datos del CSV.

    Raises:
        FileNotFoundError: Si el archivo no existe.
        ValueError: Si el archivo está vacío o tiene formato inválido.
        pd.errors.ParserError: Si el CSV está mal formateado.
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path.resolve()}")

    if path.suffix.lower() != ".csv":
        raise ValueError(f"Formato inválido. Se esperaba .csv, se recibió '{path.suffix}'")

    print(f"\n{'📂'} Cargando archivo: {path.name}")

    try:
        df = pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        # Intento secundario con latin-1 (común en archivos latinoamericanos)
        df = pd.read_csv(path, encoding="latin-1")
    except pd.errors.EmptyDataError:
        raise ValueError("El archivo CSV está vacío.")
    except pd.errors.ParserError as exc:
        raise pd.errors.ParserError(f"Error al parsear el CSV: {exc}") from exc

    if df.empty:
        raise ValueError("El DataFrame resultante está vacío.")

    print(f"{'✅'} Archivo cargado exitosamente.\n")
    return df


# ─────────────────────────────────────────────
# MÓDULO: RESUMEN GENERAL
# ─────────────────────────────────────────────

def print_overview(df: pd.DataFrame) -> dict:
    """
    Imprime y retorna el resumen estructural del DataFrame.

    Args:
        df: DataFrame analizado.

    Returns:
        Diccionario con métricas de estructura.
    """
    numeric_cols  = df.select_dtypes(include="number").columns.tolist()
    cat_cols      = df.select_dtypes(include="object").columns.tolist()
    datetime_cols = df.select_dtypes(include="datetime").columns.tolist()

    overview = {
        "filas"              : len(df),
        "columnas"           : len(df.columns),
        "columnas_numericas" : len(numeric_cols),
        "columnas_categoricas": len(cat_cols),
        "columnas_datetime"  : len(datetime_cols),
        "nombres_columnas"   : df.columns.tolist(),
        "tipos_de_datos"     : df.dtypes.to_dict(),
    }

    print(SECTION_SEP)
    print("  📊  RESUMEN GENERAL DEL DATASET")
    print(SECTION_SEP)
    print(f"  {'Filas:':<30} {overview['filas']:,}")
    print(f"  {'Columnas:':<30} {overview['columnas']}")
    print(f"  {'Columnas numéricas:':<30} {overview['columnas_numericas']}")
    print(f"  {'Columnas categóricas:':<30} {overview['columnas_categoricas']}")
    print(f"  {'Columnas fecha/hora:':<30} {overview['columnas_datetime']}")
    print()
    print("  Columnas detectadas:")
    for col, dtype in df.dtypes.items():
        tag = _dtype_tag(str(dtype))
        print(f"    • {col:<30} [{tag}]")
    print()

    return overview


def _dtype_tag(dtype_str: str) -> str:
    """Convierte el dtype de pandas a una etiqueta legible."""
    if "int" in dtype_str:
        return "Entero"
    if "float" in dtype_str:
        return "Decimal"
    if "datetime" in dtype_str:
        return "Fecha"
    if "bool" in dtype_str:
        return "Booleano"
    return "Texto"


# ─────────────────────────────────────────────
# MÓDULO: CALIDAD DE DATOS
# ─────────────────────────────────────────────

def analyze_data_quality(df: pd.DataFrame) -> dict:
    """
    Detecta valores nulos y duplicados e imprime un reporte de calidad.

    Args:
        df: DataFrame analizado.

    Returns:
        Diccionario con métricas de calidad.
    """
    null_counts  = df.isnull().sum()
    null_pct     = (null_counts / len(df) * 100).round(2)
    duplicates   = df.duplicated().sum()
    total_cells  = df.size
    total_nulls  = null_counts.sum()
    completeness = round((1 - total_nulls / total_cells) * 100, 2)

    quality = {
        "nulos_por_columna"    : null_counts.to_dict(),
        "porcentaje_nulos"     : null_pct.to_dict(),
        "total_nulos"          : int(total_nulls),
        "filas_duplicadas"     : int(duplicates),
        "completitud_pct"      : completeness,
    }

    print(SECTION_SEP)
    print("  🔍  CALIDAD DE DATOS")
    print(SECTION_SEP)
    print(f"  {'Completitud del dataset:':<35} {completeness}%")
    print(f"  {'Total de valores nulos:':<35} {total_nulls:,}")
    print(f"  {'Filas duplicadas:':<35} {duplicates:,}")
    print()

    cols_con_nulos = [(col, cnt, pct)
                      for col, cnt, pct
                      in zip(null_counts.index, null_counts.values, null_pct.values)
                      if cnt > 0]

    if cols_con_nulos:
        print("  Columnas con valores nulos:")
        for col, cnt, pct in cols_con_nulos:
            barra = "█" * int(pct / 5)
            print(f"    • {col:<28} {cnt:>4} nulos  ({pct:>5.1f}%)  {barra}")
    else:
        print("  ✅ No se detectaron valores nulos.")
    print()

    return quality


# ─────────────────────────────────────────────
# MÓDULO: ESTADÍSTICAS DESCRIPTIVAS
# ─────────────────────────────────────────────

def compute_statistics(df: pd.DataFrame) -> dict:
    """
    Calcula estadísticas descriptivas completas para columnas numéricas.

    Args:
        df: DataFrame analizado.

    Returns:
        Diccionario con estadísticas por columna numérica.
    """
    numeric_df = df.select_dtypes(include="number")

    if numeric_df.empty:
        print("  ⚠️  No se encontraron columnas numéricas.\n")
        return {}

    stats_dict: dict[str, dict] = {}

    print(SECTION_SEP)
    print("  📈  ESTADÍSTICAS DESCRIPTIVAS")
    print(SECTION_SEP)

    for col in numeric_df.columns:
        serie = numeric_df[col].dropna()

        stats = {
            "promedio"        : round(float(serie.mean()),   4),
            "mediana"         : round(float(serie.median()), 4),
            "maximo"          : round(float(serie.max()),    4),
            "minimo"          : round(float(serie.min()),    4),
            "desv_estandar"   : round(float(serie.std()),    4),
            "varianza"        : round(float(serie.var()),    4),
            "q1"              : round(float(serie.quantile(0.25)), 4),
            "q3"              : round(float(serie.quantile(0.75)), 4),
            "rango_iqr"       : round(float(serie.quantile(0.75) - serie.quantile(0.25)), 4),
            "conteo_validos"  : int(serie.count()),
            "asimetria"       : round(float(serie.skew()), 4),
            "curtosis"        : round(float(serie.kurt()), 4),
        }
        stats_dict[col] = stats

        print(f"\n  Columna: {col}")
        print(f"    {'Promedio:':<22} {stats['promedio']:>15,.4f}")
        print(f"    {'Mediana:':<22} {stats['mediana']:>15,.4f}")
        print(f"    {'Máximo:':<22} {stats['maximo']:>15,.4f}")
        print(f"    {'Mínimo:':<22} {stats['minimo']:>15,.4f}")
        print(f"    {'Desv. estándar:':<22} {stats['desv_estandar']:>15,.4f}")
        print(f"    {'Q1 (25%):':<22} {stats['q1']:>15,.4f}")
        print(f"    {'Q3 (75%):':<22} {stats['q3']:>15,.4f}")
        print(f"    {'IQR:':<22} {stats['rango_iqr']:>15,.4f}")

    print()
    return stats_dict


# ─────────────────────────────────────────────
# MÓDULO: VISUALIZACIONES
# ─────────────────────────────────────────────

def generate_charts(df: pd.DataFrame, output_dir: Path) -> list[Path]:
    """
    Genera gráficos automáticos para columnas numéricas y los guarda como PNG.

    Args:
        df         : DataFrame analizado.
        output_dir : Directorio de salida para los gráficos.

    Returns:
        Lista de rutas a los archivos PNG generados.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    numeric_df  = df.select_dtypes(include="number")
    saved_paths : list[Path] = []

    if numeric_df.empty:
        print("  ⚠️  Sin columnas numéricas para graficar.\n")
        return saved_paths

    print(SECTION_SEP)
    print("  🎨  GENERANDO VISUALIZACIONES")
    print(SECTION_SEP)

    # ── 1. Histogramas individuales ──
    for col in numeric_df.columns:
        serie = numeric_df[col].dropna()
        fig, ax = _base_figure()
        ax.hist(serie, bins=15, color=PALETTE["accent"],
                edgecolor=PALETTE["primary"], alpha=0.85)
        _style_ax(ax, title=f"Distribución: {col}",
                  xlabel=col, ylabel="Frecuencia")
        path = output_dir / f"hist_{_safe_name(col)}.png"
        fig.savefig(path, dpi=150, bbox_inches="tight",
                    facecolor=PALETTE["secondary"])
        plt.close(fig)
        saved_paths.append(path)
        print(f"  ✔  hist_{_safe_name(col)}.png")

    # ── 2. Boxplots agrupados ──
    if len(numeric_df.columns) >= 2:
        fig, ax = _base_figure(figsize=(max(8, len(numeric_df.columns) * 2), 5))
        data_to_plot = [numeric_df[col].dropna().values for col in numeric_df.columns]
        bp = ax.boxplot(data_to_plot, patch_artist=True,
                        medianprops={"color": PALETTE["accent"], "linewidth": 2})
        for patch in bp["boxes"]:
            patch.set_facecolor(PALETTE["bg"])
            patch.set_alpha(0.7)
        ax.set_xticklabels(numeric_df.columns, rotation=30, ha="right",
                           color=PALETTE["white"])
        _style_ax(ax, title="Boxplot — Columnas Numéricas", ylabel="Valor")
        path = output_dir / "boxplot_numericas.png"
        fig.savefig(path, dpi=150, bbox_inches="tight",
                    facecolor=PALETTE["secondary"])
        plt.close(fig)
        saved_paths.append(path)
        print(f"  ✔  boxplot_numericas.png")

    # ── 3. Mapa de calor de correlaciones ──
    if len(numeric_df.columns) >= 2:
        corr = numeric_df.corr()
        fig, ax = _base_figure(
            figsize=(max(6, len(corr.columns)), max(5, len(corr.columns) - 1))
        )
        cmap = plt.cm.RdYlGn
        im = ax.imshow(corr.values, cmap=cmap, aspect="auto",
                       vmin=-1, vmax=1)
        ax.set_xticks(range(len(corr.columns)))
        ax.set_yticks(range(len(corr.columns)))
        ax.set_xticklabels(corr.columns, rotation=45, ha="right",
                           color=PALETTE["white"], fontsize=9)
        ax.set_yticklabels(corr.columns, color=PALETTE["white"], fontsize=9)
        for i in range(len(corr.columns)):
            for j in range(len(corr.columns)):
                val = corr.iloc[i, j]
                color = "white" if abs(val) > 0.5 else PALETTE["primary"]
                ax.text(j, i, f"{val:.2f}", ha="center", va="center",
                        color=color, fontsize=8, fontweight="bold")
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        _style_ax(ax, title="Mapa de Correlaciones")
        path = output_dir / "correlaciones.png"
        fig.savefig(path, dpi=150, bbox_inches="tight",
                    facecolor=PALETTE["secondary"])
        plt.close(fig)
        saved_paths.append(path)
        print(f"  ✔  correlaciones.png")

    # ── 4. Nulos por columna (barra horizontal) ──
    null_counts = df.isnull().sum()
    null_counts = null_counts[null_counts > 0]
    if not null_counts.empty:
        fig, ax = _base_figure(figsize=(9, max(3, len(null_counts) * 0.7)))
        bars = ax.barh(null_counts.index, null_counts.values,
                       color=PALETTE["accent"], edgecolor=PALETTE["primary"])
        for bar, val in zip(bars, null_counts.values):
            ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                    str(val), va="center", color=PALETTE["white"], fontsize=9)
        _style_ax(ax, title="Valores Nulos por Columna", xlabel="Cantidad de Nulos")
        path = output_dir / "valores_nulos.png"
        fig.savefig(path, dpi=150, bbox_inches="tight",
                    facecolor=PALETTE["secondary"])
        plt.close(fig)
        saved_paths.append(path)
        print(f"  ✔  valores_nulos.png")

    print(f"\n  Total gráficos generados: {len(saved_paths)}\n")
    return saved_paths


def _base_figure(figsize: tuple = (9, 5)):
    """Crea figura y ejes con fondo oscuro corporativo."""
    fig, ax = plt.subplots(figsize=figsize, facecolor=PALETTE["secondary"])
    ax.set_facecolor(PALETTE["primary"])
    return fig, ax


def _style_ax(ax, title="", xlabel="", ylabel=""):
    """Aplica estilos consistentes a los ejes."""
    ax.set_title(title, color=PALETTE["white"], fontsize=13,
                 fontweight="bold", pad=12)
    ax.set_xlabel(xlabel, color=PALETTE["muted"], fontsize=10)
    ax.set_ylabel(ylabel, color=PALETTE["muted"], fontsize=10)
    ax.tick_params(colors=PALETTE["white"])
    for spine in ax.spines.values():
        spine.set_edgecolor(PALETTE["muted"])
        spine.set_linewidth(0.5)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"{x:,.0f}"
    ))


def _safe_name(text: str) -> str:
    """Convierte texto a nombre de archivo seguro."""
    return "".join(c if c.isalnum() else "_" for c in text).strip("_")


# ─────────────────────────────────────────────
# MÓDULO: EXPORTAR — TXT
# ─────────────────────────────────────────────

def export_txt_report(
    df        : pd.DataFrame,
    overview  : dict,
    quality   : dict,
    stats     : dict,
    file_path : str,
    output_dir: Path,
) -> Path:
    """
    Exporta un reporte completo en formato .txt.

    Returns:
        Ruta del archivo generado.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "reporte.txt"
    now      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines: list[str] = []
    sep = "═" * 65

    def h(text: str) -> None:
        lines.extend(["", sep, f"  {text}", sep])

    lines.append(sep)
    lines.append("  CSV INSIGHT ANALYZER — REPORTE DE ANÁLISIS")
    lines.append(sep)
    lines.append(f"  Archivo analizado : {Path(file_path).name}")
    lines.append(f"  Fecha de análisis : {now}")
    lines.append(f"  Versión           : 1.0.0  |  MillesvcStudio")
    lines.append(sep)

    # ── Resumen general ──
    h("RESUMEN GENERAL")
    lines.append(f"  Filas                : {overview['filas']:,}")
    lines.append(f"  Columnas             : {overview['columnas']}")
    lines.append(f"  Columnas numéricas   : {overview['columnas_numericas']}")
    lines.append(f"  Columnas categóricas : {overview['columnas_categoricas']}")
    lines.append("")
    lines.append("  Columnas y tipos de datos:")
    for col, dtype in overview["tipos_de_datos"].items():
        lines.append(f"    • {col:<28} {_dtype_tag(str(dtype))}")

    # ── Calidad ──
    h("CALIDAD DE DATOS")
    lines.append(f"  Completitud          : {quality['completitud_pct']}%")
    lines.append(f"  Total nulos          : {quality['total_nulos']:,}")
    lines.append(f"  Filas duplicadas     : {quality['filas_duplicadas']:,}")
    lines.append("")
    nulos = {k: v for k, v in quality["nulos_por_columna"].items() if v > 0}
    if nulos:
        lines.append("  Columnas con nulos:")
        for col, cnt in nulos.items():
            pct = quality["porcentaje_nulos"][col]
            lines.append(f"    • {col:<28} {cnt:>4} ({pct:.1f}%)")
    else:
        lines.append("  Sin valores nulos detectados.")

    # ── Estadísticas ──
    h("ESTADÍSTICAS DESCRIPTIVAS")
    if stats:
        for col, s in stats.items():
            lines.append(f"\n  [ {col} ]")
            lines.append(f"    Promedio        : {s['promedio']:>15,.4f}")
            lines.append(f"    Mediana         : {s['mediana']:>15,.4f}")
            lines.append(f"    Máximo          : {s['maximo']:>15,.4f}")
            lines.append(f"    Mínimo          : {s['minimo']:>15,.4f}")
            lines.append(f"    Desv. Estándar  : {s['desv_estandar']:>15,.4f}")
            lines.append(f"    Varianza        : {s['varianza']:>15,.4f}")
            lines.append(f"    Q1 (25%)        : {s['q1']:>15,.4f}")
            lines.append(f"    Q3 (75%)        : {s['q3']:>15,.4f}")
            lines.append(f"    IQR             : {s['rango_iqr']:>15,.4f}")
            lines.append(f"    Asimetría       : {s['asimetria']:>15,.4f}")
            lines.append(f"    Curtosis        : {s['curtosis']:>15,.4f}")
    else:
        lines.append("  Sin columnas numéricas.")

    lines.append("")
    lines.append(sep)
    lines.append("  Reporte generado automáticamente por CSV Insight Analyzer")
    lines.append(sep)

    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


# ─────────────────────────────────────────────
# MÓDULO: EXPORTAR — XLSX
# ─────────────────────────────────────────────

def export_xlsx_report(
    df        : pd.DataFrame,
    overview  : dict,
    quality   : dict,
    stats     : dict,
    output_dir: Path,
) -> Path:
    """
    Exporta un reporte completo en formato .xlsx con múltiples hojas.

    Returns:
        Ruta del archivo generado.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "reporte.xlsx"
    wb       = Workbook()

    _xlsx_sheet_overview(wb, overview, quality)
    _xlsx_sheet_stats(wb, stats)
    _xlsx_sheet_data(wb, df)

    # Elimina la hoja vacía por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(out_path)
    return out_path


# ── Helpers XLSX ──

def _xlsx_header_style(primary: bool = True) -> tuple[Font, PatternFill, Alignment, Border]:
    """Devuelve estilos de encabezado para celdas XLSX."""
    bg = "1A1A2E" if primary else "16213E"
    font    = Font(bold=True, color="FFFFFF", size=11 if primary else 10)
    fill    = PatternFill("solid", fgColor=bg)
    align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side    = Side(style="thin", color="E94560")
    border  = Border(bottom=side)
    return font, fill, align, border


def _xlsx_sheet_overview(wb: Workbook, overview: dict, quality: dict) -> None:
    ws = wb.create_sheet("Resumen General")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    font_h, fill_h, align_h, border_h = _xlsx_header_style()

    def title_row(row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font  = font_h
        cell.fill  = fill_h
        cell.alignment = align_h
        cell.border = border_h
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=2)

    def data_row(row, label, value, highlight=False):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        for c in (c1, c2):
            c.font = Font(color="ECEFF4" if not highlight else "E94560",
                          bold=highlight, size=10)
            c.fill = PatternFill("solid", fgColor="0F3460")
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18

    title_row(1,  "📊  RESUMEN GENERAL DEL DATASET")
    data_row(2,   "Archivo analizado",    overview.get("_source_file", "—"))
    data_row(3,   "Fecha de análisis",    datetime.now().strftime("%Y-%m-%d %H:%M"))
    data_row(4,   "Total de filas",       overview["filas"])
    data_row(5,   "Total de columnas",    overview["columnas"])
    data_row(6,   "Columnas numéricas",   overview["columnas_numericas"])
    data_row(7,   "Columnas categóricas", overview["columnas_categoricas"])
    data_row(8,   "Columnas fecha/hora",  overview["columnas_datetime"])

    title_row(10, "🔍  CALIDAD DE DATOS")
    data_row(11,  "Completitud del dataset", f"{quality['completitud_pct']}%", highlight=True)
    data_row(12,  "Total valores nulos",     quality["total_nulos"])
    data_row(13,  "Filas duplicadas",        quality["filas_duplicadas"])

    ws.cell(row=15, column=1, value="Columna").font = font_h
    ws.cell(row=15, column=1).fill    = fill_h
    ws.cell(row=15, column=1).alignment = align_h
    ws.cell(row=15, column=2, value="Nulos (cantidad | %)").font = font_h
    ws.cell(row=15, column=2).fill    = fill_h
    ws.cell(row=15, column=2).alignment = align_h

    row = 16
    for col, cnt in quality["nulos_por_columna"].items():
        pct = quality["porcentaje_nulos"][col]
        c1  = ws.cell(row=row, column=1, value=col)
        c2  = ws.cell(row=row, column=2, value=f"{cnt} ({pct:.1f}%)")
        for c in (c1, c2):
            c.fill = PatternFill("solid", fgColor="0F3460")
            c.font = Font(color="9EB8D9", size=10)
        row += 1


def _xlsx_sheet_stats(wb: Workbook, stats: dict) -> None:
    if not stats:
        return

    ws = wb.create_sheet("Estadísticas")
    ws.sheet_view.showGridLines = False

    headers = ["Columna", "Promedio", "Mediana", "Máximo", "Mínimo",
               "Desv. Estándar", "Varianza", "Q1", "Q3", "IQR",
               "Asimetría", "Curtosis"]

    font_h, fill_h, align_h, border_h = _xlsx_header_style()

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = font_h
        cell.fill      = fill_h
        cell.alignment = align_h
        cell.border    = border_h
        ws.column_dimensions[get_column_letter(ci)].width = 16

    ws.column_dimensions["A"].width = 26

    for ri, (col, s) in enumerate(stats.items(), start=2):
        values = [col, s["promedio"], s["mediana"], s["maximo"],
                  s["minimo"], s["desv_estandar"], s["varianza"],
                  s["q1"], s["q3"], s["rango_iqr"],
                  s["asimetria"], s["curtosis"]]
        bg = "0F3460" if ri % 2 == 0 else "16213E"
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color="ECEFF4", size=10)
            cell.alignment = Alignment(horizontal="right" if ci > 1 else "left",
                                        vertical="center")


def _xlsx_sheet_data(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Datos Originales")
    ws.sheet_view.showGridLines = False

    font_h, fill_h, align_h, border_h = _xlsx_header_style(primary=False)

    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = font_h
        cell.fill      = fill_h
        cell.alignment = align_h
        cell.border    = border_h
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 4)

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        bg = "0F3460" if ri % 2 == 0 else "16213E"
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci,
                           value=None if pd.isna(val) else val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color="ECEFF4", size=9)
            cell.alignment = Alignment(vertical="center")


# ─────────────────────────────────────────────
# ORQUESTADOR PRINCIPAL
# ─────────────────────────────────────────────

def run_analysis(file_path: str | Path) -> None:
    """
    Ejecuta el pipeline completo de análisis para un archivo CSV.

    Pipeline:
        1. Carga del archivo
        2. Resumen estructural
        3. Análisis de calidad
        4. Estadísticas descriptivas
        5. Generación de gráficos
        6. Exportación de reportes (TXT + XLSX)

    Args:
        file_path: Ruta al archivo CSV a analizar.
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    charts_dir = REPORTS_DIR / "charts"
    now_str    = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("\n" + "═" * 65)
    print("  🔬  CSV INSIGHT ANALYZER  |  MillesvcStudio  v1.0.0")
    print("═" * 65)

    # ── 1. Carga ──
    df = load_csv(file_path)

    # ── 2. Resumen ──
    overview = print_overview(df)
    overview["_source_file"] = Path(file_path).name

    # ── 3. Calidad ──
    quality = analyze_data_quality(df)

    # ── 4. Estadísticas ──
    stats = compute_statistics(df)

    # ── 5. Gráficos ──
    chart_paths = generate_charts(df, charts_dir)

    # ── 6. Exportar TXT ──
    txt_path = export_txt_report(
        df, overview, quality, stats,
        str(file_path), REPORTS_DIR
    )

    # ── 6. Exportar XLSX ──
    xlsx_path = export_xlsx_report(df, overview, quality, stats, REPORTS_DIR)

    # ── Resumen final ──
    print(SECTION_SEP)
    print("  ✅  ANÁLISIS COMPLETADO")
    print(SECTION_SEP)
    print(f"  {'Reporte TXT:':<32} {txt_path}")
    print(f"  {'Reporte XLSX:':<32} {xlsx_path}")
    print(f"  {'Gráficos PNG:':<32} {charts_dir}  ({len(chart_paths)} archivos)")
    print()
    print("  Para visualizar los reportes, abre la carpeta /reports/")
    print(SECTION_SEP + "\n")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    # Acepta ruta por argumento o usa el CSV de muestra por defecto
    csv_path = sys.argv[1] if len(sys.argv) > 1 else DATA_DIR / "sample.csv"

    try:
        run_analysis(csv_path)
    except FileNotFoundError as e:
        print(f"\n❌  Error: {e}")
        print("   Verifica que el archivo exista y la ruta sea correcta.\n")
        sys.exit(1)
    except ValueError as e:
        print(f"\n❌  Error de datos: {e}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌  Error inesperado: {e}\n")
        sys.exit(1)
